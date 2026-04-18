"""Excel import/export routes."""
import io
from datetime import date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backend.database import get_db
from backend.models import TimeEntry, Project

router = APIRouter(prefix="/api/excel", tags=["excel"])

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


@router.get("/export")
def export_excel(
    worker_name: str = None,
    project_id: int = None,
    start_date: date = None,
    end_date: date = None,
    db: Session = Depends(get_db),
):
    q = db.query(TimeEntry).options(joinedload(TimeEntry.project))
    if worker_name:
        q = q.filter(TimeEntry.worker_name == worker_name)
    if project_id:
        q = q.filter(TimeEntry.project_id == project_id)
    if start_date:
        q = q.filter(TimeEntry.date >= start_date)
    if end_date:
        q = q.filter(TimeEntry.date <= end_date)
    entries = q.order_by(TimeEntry.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    headers = ["Date", "Worker", "Project", "Location", "Hours", "Overtime", "Task", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, e in enumerate(entries, 2):
        vals = [
            e.date.isoformat(), e.worker_name, e.project.name,
            e.project.location, e.hours, e.overtime,
            e.task_description, e.notes,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = THIN_BORDER

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=timesheet_export.xlsx"},
    )


@router.post("/import")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx files are supported")

    wb = load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    errors = []

    for i, row in enumerate(rows, 2):
        if not row or len(row) < 5:
            continue
        try:
            entry_date = row[0] if isinstance(row[0], date) else date.fromisoformat(str(row[0]))
            worker = str(row[1]).strip()
            project_name = str(row[2]).strip()
            hours = float(row[4])
            overtime = float(row[5]) if len(row) > 5 and row[5] else 0.0
            task = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            notes = str(row[7]).strip() if len(row) > 7 and row[7] else ""

            project = db.query(Project).filter(Project.name == project_name).first()
            if not project:
                location = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                project = Project(name=project_name, location=location)
                db.add(project)
                db.flush()

            entry = TimeEntry(
                worker_name=worker, date=entry_date, hours=hours,
                project_id=project.id, task_description=task,
                overtime=overtime, notes=notes,
            )
            db.add(entry)
            imported += 1
        except Exception as ex:
            errors.append(f"Row {i}: {str(ex)}")

    db.commit()
    return {"imported": imported, "errors": errors}
